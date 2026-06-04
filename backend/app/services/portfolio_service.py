"""
Servicio de lectura del portafolio de soluciones desde Excel.
Lee ListaPortafolio.xlsx y expone los productos disponibles.
"""
from dataclasses import dataclass, field
from typing import List, Optional
from pathlib import Path

import openpyxl

# Scheme types available in the MVP (string values matching SchemeType enum)
MVP_SCHEME_STRINGS: List[str] = ["licensing", "services", "support_maintenance"]


@dataclass
class PortfolioProduct:
    """Producto/servicio del portafolio de Quipux."""
    name: str
    product_type: str  # "Plataforma" o "Servicio QloudSI"
    description: str
    business_framework: str
    revenue_info: str
    operational_costs: str
    monetization_model: str
    pricing_model: str
    country: str
    # Categoría opcional proveniente de la propuesta (ej: "nuevo", "modernización")
    category: str = ""
    # Comma-separated scheme types allowed for this product (col 9 in Excel).
    # Empty list means all MVP schemes are allowed.
    allowed_schemes: List[str] = field(default_factory=list)


class PortfolioNotFoundError(Exception):
    """Error cuando no se encuentra el archivo de portafolio."""
    pass


class PortfolioService:
    """Servicio para cargar y consultar el portafolio de soluciones."""

    # Configuración de lectura del Excel (fila de headers y columnas)
    HEADER_ROW = 9
    DATA_START_ROW = 10
    SHEET_NAME = "Hoja2"

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self._products: Optional[List[PortfolioProduct]] = None

    def _validate_file(self) -> None:
        """Valida que el archivo de portafolio existe."""
        # Resolver ruta relativa desde la ubicación de este script
        if not self.file_path.is_absolute():
            # Si es relativa, resolverla desde la carpeta backend
            service_dir = Path(__file__).parent.parent.parent
            resolved_path = service_dir / self.file_path
        else:
            resolved_path = self.file_path

        if not resolved_path.exists():
            raise PortfolioNotFoundError(
                f"Archivo de portafolio no encontrado: {resolved_path}\n"
                f"Ruta original: {self.file_path}\n"
                f"Buscó en: {resolved_path}"
            )

        # Actualizar la ruta interna a la resuelta
        self.file_path = resolved_path

    def load_products(self) -> List[PortfolioProduct]:
        """Carga todos los productos del portafolio desde el Excel."""
        self._validate_file()

        wb = openpyxl.load_workbook(str(self.file_path), read_only=True)
        ws = wb[self.SHEET_NAME]

        products: List[PortfolioProduct] = []

        for row in ws.iter_rows(
            min_row=self.DATA_START_ROW,
            max_row=ws.max_row,
            values_only=True
        ):
            name = row[0]
            if not name:
                continue

            # Column 9: allowed_schemes — comma-separated scheme types.
            # Empty or missing → all MVP schemes are allowed.
            raw_schemes = row[9] if len(row) > 9 else None
            if raw_schemes:
                parsed = [s.strip() for s in str(raw_schemes).split(",") if s.strip()]
                allowed_schemes = [s for s in parsed if s in MVP_SCHEME_STRINGS]
            else:
                allowed_schemes = []

            product = PortfolioProduct(
                name=str(name).strip(),
                product_type=str(row[1] or "").strip(),
                description=str(row[2] or "").strip(),
                business_framework=str(row[3] or "").strip(),
                revenue_info=str(row[4] or "").strip(),
                operational_costs=str(row[5] or "").strip(),
                monetization_model=str(row[6] or "").strip(),
                pricing_model=str(row[7] or "").strip(),
                country=str(row[8] or "").strip(),
                allowed_schemes=allowed_schemes,
            )
            products.append(product)

        wb.close()
        self._products = products
        return products

    def get_products(self) -> List[PortfolioProduct]:
        """Obtiene la lista de productos (carga si no se ha cargado)."""
        if self._products is None:
            self.load_products()
        return self._products  # type: ignore

    def search_products(self, query: str) -> List[PortfolioProduct]:
        """Busca productos por nombre (case-insensitive)."""
        products = self.get_products()
        query_lower = query.lower()
        return [p for p in products if query_lower in p.name.lower()]

    def filter_by_type(self, product_type: str) -> List[PortfolioProduct]:
        """Filtra productos por tipo (Plataforma, Servicio QloudSI)."""
        products = self.get_products()
        return [p for p in products if product_type.lower() in p.product_type.lower()]

    def get_by_names(self, names: List[str]) -> List[PortfolioProduct]:
        """Obtiene productos específicos por sus nombres exactos."""
        products = self.get_products()
        name_set = {n.lower() for n in names}
        return [p for p in products if p.name.lower() in name_set]

    def get_allowed_schemes_for_products(self, product_names: List[str]) -> List[str]:
        """Returns the intersection of allowed schemes across the given product names.

        A product with no scheme restrictions (empty allowed_schemes or not found)
        contributes all MVP schemes to the intersection — it does not restrict anything.
        Returns an empty list when product_names is empty.
        An empty result means the selected products have no schemes in common.
        """
        if not product_names:
            return []

        all_products = self.get_products()
        product_map = {p.name.lower(): p for p in all_products}

        result: set = set(MVP_SCHEME_STRINGS)
        for name in product_names:
            product = product_map.get(name.lower())
            if product is None or not product.allowed_schemes:
                # No restriction → contributes all MVP schemes; does not narrow the set
                product_schemes = set(MVP_SCHEME_STRINGS)
            else:
                product_schemes = set(product.allowed_schemes)
            result &= product_schemes
            if not result:
                return []

        return sorted(result, key=lambda s: MVP_SCHEME_STRINGS.index(s) if s in MVP_SCHEME_STRINGS else 999)
